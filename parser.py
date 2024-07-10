# =================== imports ===================
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import configparser
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import Alignment

print('===================================')
print('Программа парсинга MegaMarket!')
print('Не закрывайте программу до окончания её работы!')
print('===================================')

while True:
    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    # Add headers to the worksheet
    worksheet.append(["Страница", "Название", "Цена ₽", "Кол-во бонусов", "Процент"])
    # Переменные
    PATH_chrome_driver = f'src\\chromedriver.exe'
    link = input('Введите ссылку: ')
    countPage = input('Введите кол-во страниц для парсинга: ')
    file_name = input('Введите только название файла, где будет сохранён парсинг: ')
    counter = 0
    print(f'Открытие браузера...')
    service = Service(executable_path=PATH_chrome_driver)
    service.start()
    options = Options()
    browser = webdriver.Remote(service.service_url, options=options)
    for page in range(1, int(countPage) + 1):
        browser.get(f'{link}/page-{page}/')
        try:
            search_elements = browser.find_elements(By.CSS_SELECTOR, '.item-info')

            for element in search_elements:
                if element.find_elements(By.CSS_SELECTOR, '.bonus-amount'):
                    linkProduct = element.find_element(By.CSS_SELECTOR, '.item-title a').get_attribute('href')
                    name = element.find_element(By.CSS_SELECTOR, '.item-title').text
                    price = element.find_element(By.CSS_SELECTOR, '.item-price').text
                    price = price.replace(' ', '').replace('₽', '')
                    bonus_amount = element.find_element(By.CSS_SELECTOR, '.bonus-amount').text
                    bonus_amount = bonus_amount.replace(' ', '').replace(' ', '')
                    procent = (int(bonus_amount) / int(price)) * 100
                    procent = format(procent, '.0f')

                    worksheet.append([page, f'=HYPERLINK("{linkProduct}", "{name}")', int(price), int(bonus_amount), int(procent)])
                    print(f'Страница: {page}\t Наименование: {name}\t Цена: {price}\t Кол-во бонусов: {bonus_amount}\t Процент: {procent} %')
                    counter += 1
        except NoSuchElementException:
            print(f'На странице {page} нет необходимых элементов. Скрипт завершается.')
            break

    # Set alignment to center and adjust column width
    worksheet.column_dimensions['A'].width = 15 * 0.7
    worksheet.column_dimensions['B'].width = 100 * 0.7
    worksheet.column_dimensions['C'].width = 25 * 0.7
    worksheet.column_dimensions['D'].width = 25 * 0.7
    worksheet.column_dimensions['E'].width = 20 * 0.7

    # Create a new font object with bold style
    bold_font = Font(bold=True)

    # Set the font of cells in the range of A1:E1 to bold
    worksheet['A1'].font = bold_font
    worksheet['B1'].font = bold_font
    worksheet['C1'].font = bold_font
    worksheet['D1'].font = bold_font
    worksheet['E1'].font = bold_font
    worksheet['A1'].fill = PatternFill(start_color='50c878', end_color='50c878', fill_type='solid')
    worksheet['B1'].fill = PatternFill(start_color='50c878', end_color='50c878', fill_type='solid')
    worksheet['C1'].fill = PatternFill(start_color='50c878', end_color='50c878', fill_type='solid')
    worksheet['D1'].fill = PatternFill(start_color='50c878', end_color='50c878', fill_type='solid')
    worksheet['E1'].fill = PatternFill(start_color='50c878', end_color='50c878', fill_type='solid')
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet['B'][1:]:
        if cell.value:
            # Set the cell's font to blue and underline
            cell.font = openpyxl.styles.Font(color='0000FF', underline='single')

    # Set alignment to center for all cells in the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print('Программа успешно отработала!')
    print(f'Отчёт сохранён в файле {file_name}.xlsx')
    print(f'Кол-во товаров с бонусами:  {counter}')
    # Save the workbook to a file
    workbook.save(f'result\\{file_name}.xlsx')
    # Close the browser
    browser.quit()
